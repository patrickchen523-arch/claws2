#!/usr/bin/env python3
"""读取 pm_schedule.xlsx，输出今天开始 + 近3天内需处理的任务提醒"""

import sys
from datetime import date, timedelta

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed")
    sys.exit(1)

EXCEL_PATH = "/root/.openclaw/workspace/pm_schedule.xlsx"

def main():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
    except FileNotFoundError:
        print(f"找不到文件：{EXCEL_PATH}")
        sys.exit(1)

    ws = wb.active
    today = date.today()
    deadline = today + timedelta(days=3)

    # 读取表头，找列索引
    headers = [cell.value for cell in ws[1]]
    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    idx_name     = col("任务名称")
    idx_owner    = col("负责人")
    idx_start    = col("开始日期")
    idx_due      = col("截止日期")
    idx_priority = col("优先级")
    idx_status   = col("状态")

    starting_today = []
    due_soon = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        name     = row[idx_name]     if idx_name is not None else ""
        owner    = row[idx_owner]    if idx_owner is not None else ""
        start    = row[idx_start]    if idx_start is not None else None
        due      = row[idx_due]      if idx_due is not None else None
        priority = row[idx_priority] if idx_priority is not None else ""
        status   = row[idx_status]   if idx_status is not None else ""

        if not name:
            continue

        # 跳过已完成
        if status and "完成" in str(status):
            continue

        # 解析日期
        def parse_date(d):
            if d is None:
                return None
            from datetime import datetime as dt
            if isinstance(d, dt):
                return d.date()
            if isinstance(d, date):
                return d
            try:
                return dt.strptime(str(d).strip(), "%Y-%m-%d").date()
            except:
                return None

        start_date = parse_date(start)
        due_date   = parse_date(due)

        item = {
            "name": name,
            "owner": owner or "-",
            "start": str(start_date) if start_date else "-",
            "due": str(due_date) if due_date else "-",
            "priority": priority or "-",
            "status": status or "-",
        }

        # 今天开始的任务
        if start_date == today:
            starting_today.append(item)
        # 近3天内截止（不含今天开始的，避免重复）
        elif due_date and today <= due_date <= deadline:
            due_soon.append(item)

    # 构建输出
    lines = [f"📋 项目进度提醒 · {today.strftime('%Y-%m-%d')}"]
    lines.append("")

    if starting_today:
        lines.append("🟢 今日开始")
        for t in starting_today:
            lines.append(f"  • {t['name']}  负责人：{t['owner']}  截止：{t['due']}  优先级：{t['priority']}")
    else:
        lines.append("🟢 今日开始：无")

    lines.append("")

    if due_soon:
        lines.append("⚠️ 近3天内截止")
        for t in due_soon:
            lines.append(f"  • {t['name']}  负责人：{t['owner']}  截止：{t['due']}  优先级：{t['priority']}  状态：{t['status']}")
    else:
        lines.append("⚠️ 近3天内截止：无")

    print("\n".join(lines))

if __name__ == "__main__":
    main()
